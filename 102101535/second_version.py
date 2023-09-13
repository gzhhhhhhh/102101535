import collections
import json
import requests
import re
import time
import openpyxl
import pandas as pd


def get_bvid(page_number, number):  # 根据搜索页api获取视频的bvid号

    # url即网址
    url = 'https://api.bilibili.com/x/web-interface/search/all/v2?' \
          'page='+str(page_number)+'&keyword=%E6%97%A5%E6%9C%AC%E6%A0%B8%E6%B1%A1%E6%9F%93%E6%B0%B4%E6%8E%92%E6%B5%B7'

    headers = {

        'cookie': 'buvid3=AF7E4336-683D-2AEB-5667-5FB21535E67187193infoc; i-wanna-go-back=-1;'
                  ' _uuid=F93E3C87-10382-23AD-824B-27A193361E8E66199infoc; FEED_LIVE_VERSION=V8;'
                  ' header_theme_version=CLOSE; buvid_fp=2ba3ad7f6acb16191f1ca49ffc5c66b3; '
                  'home_feed_column=5; nostalgia_conf=-1; CURRENT_FNVAL=4048; '
                  'rpdid=|(u))kkYu|Ru0J\'uY))m~YRl); DedeUserID=527426084; '
                  'DedeUserID__ckMd5=00b587f1778e1e7b; b_ut=5; '
                  'LIVE_BUVID=AUTO5716894257275090; '
                  'CURRENT_QUALITY=80; SESSDATA=430f2148%2C1705238175%2C979d2%2A72x'
                  'oe6cwW6v8CmqC2s1IXhkMwzPMZyCx5XdhwFHZX-XN4sTGTJxXrWqVh2aK1s569jWiUxbQAAVAA; '
                  'bili_jct=40f034f73101fa664a3d63ef89673470; sid=6wb18330; '
                  'b_nut=1689748974; buvid4=7A0E5B98-695A-4A87-3965-0'
                  'F4C57AE23D188249-023072016-Mk4wjKcJQ46e8fG6nCNzkDivye7Q'
                  'm5pptEQXu4qFguaYYVRWkzj5LQ%3D%3D; PVID=1; browser_resolution='
                  '1528-750; bili_ticket=eyJhbGciOiJIUzI1NiIsImtpZCI6InMwMyIsInR5cCI6IkpX'
                  'VCJ9.eyJleHAiOjE2OTQyNDQ0NTIsImlhdCI6MTY5Mzk4NTI1MiwicGx0IjotMX0.hSJQoBvDe4fdw9g'
                  'OlHh_94WTswnSf8dKbaUs8rjfyfk; bili_ticket_expires=1694244452; b_lsid=75EC3556_18A69B52964',

        'origin': 'https://www.bilibili.com',

        'referer': 'https://www.bilibili.com/video/BV1yF411C7ZJ/?spm_id_from='
                   '333.337.search-card.all.click&vd_source=e5ea948412c2a8820992ad19400de8ab',

        'user-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.'
                      '6 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36 Edg/116.0.1938.69'

    }   # 加上我的头部信息，伪装请求来源，减少被拦截的可能性

    response = requests.get(url=url, headers=headers).text      # 利用HTTP的get方法，以json格式得到文件
    json_data = json.loads(response)        # 用python内置函数将json文件转换成字典格式
    return json_data['data']['result'][11]['data'][number]['bvid']      # 根据key得到value


def get_cid(bvid):      # 视频的bvid号得到指定视频的cid号

    url = 'https://api.bilibili.com/x/player/pagelist?bvid='+str(bvid)+'&jsonp=jsonp'

    response = requests.get(url).text    # 利用HTTP的get方法，以json格式得到文件

    json_dict = json.loads(response)     # 用python内置函数将json文件转换成字典格式

    return json_dict['data'][0]['cid']      # 根据key得到value


def get_data():

    workbook = openpyxl.Workbook()       # 创建工作簿
    sheet = workbook.active     # 创建工作表

    headers = {

        'cookie': 'buvid3=AF7E4336-683D-2AEB-5667-5FB21535E67187193infoc; i-wanna-go-back=-1;'
                  ' _uuid=F93E3C87-10382-23AD-824B-27A193361E8E66199infoc; FEED_LIVE_VERSION=V8;'
                  ' header_theme_version=CLOSE; buvid_fp=2ba3ad7f6acb16191f1ca49ffc5c66b3; '
                  'home_feed_column=5; nostalgia_conf=-1; CURRENT_FNVAL=4048; '
                  'rpdid=|(u))kkYu|Ru0J\'uY))m~YRl); DedeUserID=527426084; '
                  'DedeUserID__ckMd5=00b587f1778e1e7b; b_ut=5; '
                  'LIVE_BUVID=AUTO5716894257275090; '
                  'CURRENT_QUALITY=80; SESSDATA=430f2148%2C1705238175%2C979d2%2A72x'
                  'oe6cwW6v8CmqC2s1IXhkMwzPMZyCx5XdhwFHZX-XN4sTGTJxXrWqVh2aK1s569jWiUxbQAAVAA; '
                  'bili_jct=40f034f73101fa664a3d63ef89673470; sid=6wb18330; '
                  'b_nut=1689748974; buvid4=7A0E5B98-695A-4A87-3965-0'
                  'F4C57AE23D188249-023072016-Mk4wjKcJQ46e8fG6nCNzkDivye7Q'
                  'm5pptEQXu4qFguaYYVRWkzj5LQ%3D%3D; PVID=1; browser_resolution='
                  '1528-750; bili_ticket=eyJhbGciOiJIUzI1NiIsImtpZCI6InMwMyIsInR5cCI6IkpX'
                  'VCJ9.eyJleHAiOjE2OTQyNDQ0NTIsImlhdCI6MTY5Mzk4NTI1MiwicGx0IjotMX0.hSJQoBvDe4fdw9g'
                  'OlHh_94WTswnSf8dKbaUs8rjfyfk; bili_ticket_expires=1694244452; b_lsid=75EC3556_18A69B52964',

        'origin': 'https://www.bilibili.com',

        'referer': 'https://www.bilibili.com/video/BV1yF411C7ZJ/?spm_id_from='
                   '333.337.search-card.all.click&vd_source=e5ea948412c2a8820992ad19400de8ab',

        'user-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.'
                      '6 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36 Edg/116.0.1938.69'

    }       # 加上我的头部信息，伪装请求来源，减少被拦截的可能性

    sheet.append(['弹幕'])    # 为表格增加列名

    for i in range(1, 16):  # 由于获取bvid的网址是一页20个，所以外部循环15次，内部循环20次
        for j in range(20):
            url = f'https://api.bilibili.com/x/v2/dm/web' \
                  f'/history/seg.so?type=1&oid={get_cid(get_bvid(i,j))}&date=2023-09-12'

            time.sleep(1)
            response = requests.get(url=url, headers=headers)   # 利用HTTP的get方法，得到json格式文件

            data = re.findall('.*?([\u4e00-\u9fa5]+).*', response.text)     # re.findall第一个参数pattern是模式串，第二个是字符串
            # 以list形式返回符合模式串格式的所有字符串

            for index in data:
                print(index)
                sheet.append([index])   # 添加到所创建的工作表sheet中
    workbook.save('全部弹幕.xlsx')    # 保存工作簿


def calculate_frequency():

    workbook = openpyxl.Workbook()  # 创建一个工作簿
    sheet = workbook.active  # 在工作簿中创建一个工作表
    sheet.append(['弹幕'])  # 为表加上列名
    sheet.cell(row=1, column=2).value = '频次'

    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active
    new_sheet.append(['弹幕'])
    new_sheet.cell(row=1, column=2).value = '频次'

    excel_path = '全部弹幕.xlsx'
    fd = pd.read_excel(excel_path)  # 读出文件
    lines = fd['弹幕']    # 得到列名为弹幕的一列数据
    text = ' '.join(lines.astype(str))  # 先把得到的数据转换成str类型文件，在用空格把它们连接起来
    words = text.split()    # 将得到的字符串分割成列表
    word_counts = collections.Counter(words)    # 用collections库中的Counter类统计每个词出现的次数

    sorted_word_counts = sorted(word_counts.items(), key=lambda x: x[1], reverse=True)  # 按第二列对词频统计结果按照频次进行排序

    for word, count in sorted_word_counts:
        sheet.append([word, count])

    for i in range(20):
        new_sheet.append(sorted_word_counts[i])
        print("排名第" + str(i+1) + ": ", end="")
        print(sorted_word_counts[i])

    workbook.save('统计弹幕出现次数.xlsx')
    new_workbook.save('排名前20的弹幕数量.xlsx')


def main():
    get_data()
    calculate_frequency()


if __name__ == '__main__':
    main()

